# backend_mongo.py
"""
FastAPI + MongoDB (motor) backend for row-centric spreadsheet.

Run:
    pip install fastapi uvicorn motor pydantic openpyxl python-multipart python-dateutil
    MONGODB_URI=mongodb://localhost:27017 uvicorn backend_mongo:app --reload

Notes:
- For production, set MONGODB_URI env var, enable authentication, and use Redis pub/sub
  if you run multiple backend instances (so real-time messages are delivered across instances).
"""
import os
import io
import uuid
import json
from datetime import datetime
from typing import Dict, Any, Optional, List

from fastapi import FastAPI, WebSocket, WebSocketDisconnect, UploadFile, File, HTTPException, Body
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field
import motor.motor_asyncio
import openpyxl
from dateutil import parser as dateparser
import asyncio

MONGO_URI = os.getenv("MONGODB_URI", "mongodb://localhost:27017")
DB_NAME = os.getenv("MONGODB_DB", "sheets_db")

client = motor.motor_asyncio.AsyncIOMotorClient(MONGO_URI)
db = client[DB_NAME]

app = FastAPI(title="Row-centric spreadsheet (FastAPI + MongoDB)")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ---------- Pydantic models ----------
class CellModel(BaseModel):
    type: str
    value: Any = None
    meta: Optional[Dict[str, Any]] = None

class RowModel(BaseModel):
    row_index: int
    cells: Dict[str, CellModel]
    version: Optional[int] = None

class CreateSpreadsheetReq(BaseModel):
    title: str

class PatchRowReq(BaseModel):
    changes: Dict[str, Any]  # e.g. {"B": {"old": {...}, "new": {...}}}
    user_id: Optional[str] = None
    expected_version: Optional[int] = None

# ---------- Utilities ----------
def now_iso():
    return datetime.utcnow().isoformat()

def parse_typed_value(value):
    """Try to detect number/date/datetime; otherwise return string type."""
    if value is None:
        return {"type": "string", "value": None}
    if isinstance(value, (int, float, bool)):
        return {"type": "number", "value": value}
    if isinstance(value, dict):
        return {"type": "json", "value": value}
    s = str(value).strip()
    # number
    try:
        if '.' in s:
            f = float(s)
            return {"type": "number", "value": f}
        else:
            i = int(s)
            return {"type": "number", "value": i}
    except Exception:
        pass
    # datetime
    try:
        dt = dateparser.parse(s)
        # decide date vs datetime naive heuristic
        if dt.hour == 0 and dt.minute == 0 and dt.second == 0 and ':' not in s:
            return {"type": "date", "value": dt.date().isoformat()}
        else:
            return {"type": "datetime", "value": dt.isoformat()}
    except Exception:
        pass
    return {"type": "string", "value": s}

# ---------- Simple in-memory WebSocket connection manager ----------
# Map: sheet_id -> set of WebSocket
class ConnectionManager:
    def __init__(self):
        self._lock = asyncio.Lock()
        self.connections: Dict[str, List[WebSocket]] = {}

    async def connect(self, sheet_id: str, websocket: WebSocket):
        await websocket.accept()
        async with self._lock:
            self.connections.setdefault(sheet_id, []).append(websocket)

    async def disconnect(self, sheet_id: str, websocket: WebSocket):
        async with self._lock:
            conns = self.connections.get(sheet_id)
            if not conns:
                return
            if websocket in conns:
                conns.remove(websocket)
            if len(conns) == 0:
                self.connections.pop(sheet_id, None)

    async def broadcast(self, sheet_id: str, message: Dict[str, Any]):
        async with self._lock:
            conns = list(self.connections.get(sheet_id, []))
        coros = []
        for ws in conns:
            coros.append(ws.send_json(message))
        if coros:
            await asyncio.gather(*coros, return_exceptions=True)

manager = ConnectionManager()

# ---------- Endpoints ----------
@app.post("/api/spreadsheets")
async def create_spreadsheet(req: CreateSpreadsheetReq):
    doc = {"_id": str(uuid.uuid4()), "title": req.title, "created_at": now_iso(), "updated_at": now_iso()}
    await db.spreadsheets.insert_one(doc)
    return {"id": doc["_id"], "title": req.title}

@app.get("/api/spreadsheets/{sheet_id}/rows")
async def get_rows(sheet_id: str, start: int = 1, end: int = 100):
    cursor = db.rows.find({"spreadsheet_id": sheet_id, "row_index": {"$gte": start, "$lte": end}}).sort("row_index", 1)
    rows = []
    async for r in cursor:
        r["_id"] = str(r["_id"])
        rows.append({"row_index": r["row_index"], "cells": r.get("cells", {}), "version": r.get("version", 1)})
    return rows

@app.post("/api/spreadsheets/{sheet_id}/rows")
async def insert_row(sheet_id: str, payload: RowModel):
    # ensure sheet exists
    sheet = await db.spreadsheets.find_one({"_id": sheet_id})
    if not sheet:
        raise HTTPException(status_code=404, detail="spreadsheet_not_found")
    existing = await db.rows.find_one({"spreadsheet_id": sheet_id, "row_index": payload.row_index})
    if existing:
        raise HTTPException(status_code=400, detail="row_exists")
    cells_norm = {}
    for col, c in payload.cells.items():
        if isinstance(c, dict):
            cells_norm[col] = c
        else:
            cells_norm[col] = parse_typed_value(c)
    row_doc = {
        "_id": str(uuid.uuid4()),
        "spreadsheet_id": sheet_id,
        "row_index": payload.row_index,
        "cells": cells_norm,
        "version": 1,
        "updated_by": None,
        "updated_at": now_iso(),
    }
    await db.rows.insert_one(row_doc)
    # audit change
    change = {
        "_id": str(uuid.uuid4()),
        "spreadsheet_id": sheet_id,
        "row_id": row_doc["_id"],
        "user_id": None,
        "op_type": "insert_row",
        "payload": {"row_index": payload.row_index, "cells": cells_norm},
        "created_at": now_iso(),
    }
    await db.changes.insert_one(change)
    # broadcast
    await manager.broadcast(sheet_id, {"type": "row_inserted", "row": {"row_index": row_doc["row_index"], "cells": cells_norm, "version": 1}})
    return {"ok": True, "row_index": row_doc["row_index"]}

@app.patch("/api/spreadsheets/{sheet_id}/rows/{row_index}")
async def patch_row(sheet_id: str, row_index: int, req: PatchRowReq = Body(...)):
    row = await db.rows.find_one({"spreadsheet_id": sheet_id, "row_index": row_index})
    if not row:
        raise HTTPException(status_code=404, detail="row_not_found")
    current_version = row.get("version", 1)
    if req.expected_version is not None and req.expected_version != current_version:
        raise HTTPException(status_code=409, detail={"error": "version_mismatch", "current_version": current_version})
    cells = dict(row.get("cells", {}))
    changes_record = {}
    for col, change in req.changes.items():
        new_cell = change.get("new")
        if new_cell is None:
            old = cells.pop(col, None)
            changes_record[col] = {"old": old, "new": None}
        else:
            if not isinstance(new_cell, dict):
                normalized = parse_typed_value(new_cell)
            else:
                normalized = new_cell
            old = cells.get(col)
            cells[col] = normalized
            changes_record[col] = {"old": old, "new": normalized}
    new_version = current_version + 1
    update_doc = {
        "$set": {
            "cells": cells,
            "version": new_version,
            "updated_by": req.user_id,
            "updated_at": now_iso()
        }
    }
    await db.rows.update_one({"_id": row["_id"]}, update_doc)
    # audit
    change = {
        "_id": str(uuid.uuid4()),
        "spreadsheet_id": sheet_id,
        "row_id": row["_id"],
        "user_id": req.user_id,
        "op_type": "update_cells",
        "payload": {"changes": changes_record, "prev_version": current_version, "new_version": new_version},
        "created_at": now_iso(),
    }
    await db.changes.insert_one(change)
    # broadcast
    await manager.broadcast(sheet_id, {"type": "row_updated", "row": {"row_index": row_index, "cells": cells, "version": new_version}})
    return {"ok": True, "row_index": row_index, "version": new_version}

@app.get("/api/spreadsheets/{sheet_id}/rows/{row_index}/history")
async def row_history(sheet_id: str, row_index: int, limit: int = 50):
    row = await db.rows.find_one({"spreadsheet_id": sheet_id, "row_index": row_index})
    if not row:
        raise HTTPException(status_code=404, detail="row_not_found")
    cursor = db.changes.find({"spreadsheet_id": sheet_id, "row_id": row["_id"]}).sort("created_at", -1).limit(limit)
    out = []
    async for c in cursor:
        out.append({
            "id": c["_id"],
            "op_type": c["op_type"],
            "payload": c.get("payload"),
            "created_at": c.get("created_at"),
            "user_id": c.get("user_id")
        })
    return out

# WebSocket endpoint for realtime collaboration
@app.websocket("/ws/{sheet_id}")
async def websocket_endpoint(websocket: WebSocket, sheet_id: str):
    """
    Simple protocol:
    - Client subscribes by opening ws to /ws/{sheet_id}

    - Client -> Server messages (JSON):
      { "type": "update", "row_index": 5, "changes": {"B": {"old":..., "new":...}}, "user_id": "u1", "expected_version": 3 }
      { "type": "ping" }

    - Server -> Client broadcasts:
      { "type": "row_updated", "row": {...} }
      { "type": "row_inserted", "row": {...} }
      { "type": "ack", "result": {...} }  # ack for websocket-sent update
    """
    await manager.connect(sheet_id, websocket)
    try:
        while True:
            data = await websocket.receive_json()
            t = data.get("type")
            if t == "update":
                # Reuse patch_row logic but call directly
                try:
                    req = PatchRowReq(changes=data.get("changes", {}), user_id=data.get("user_id"), expected_version=data.get("expected_version"))
                    result = await patch_row(sheet_id, data.get("row_index"), req)
                    await websocket.send_json({"type": "ack", "result": result})
                except HTTPException as e:
                    await websocket.send_json({"type": "error", "detail": getattr(e, "detail", str(e))})
            elif t == "ping":
                await websocket.send_json({"type": "pong"})
            else:
                await websocket.send_json({"type": "unknown"})
    except WebSocketDisconnect:
        await manager.disconnect(sheet_id, websocket)
    except Exception:
        # ensure we remove on unexpected errors
        await manager.disconnect(sheet_id, websocket)

# Import XLSX
@app.post("/api/spreadsheets/import_xlsx")
async def import_xlsx(file: UploadFile = File(...), title: Optional[str] = None):
    content = await file.read()
    wb = openpyxl.load_workbook(filename=io.BytesIO(content), data_only=False)
    sheet = wb.active
    sheet_doc = {"_id": str(uuid.uuid4()), "title": title or file.filename, "created_at": now_iso(), "updated_at": now_iso()}
    await db.spreadsheets.insert_one(sheet_doc)
    rows_created = 0
    bulk = []
    for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        cells = {}
        for col_idx, cell_value in enumerate(row):
            if cell_value is None:
                continue
            col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
            cells[col_letter] = parse_typed_value(cell_value)
        row_doc = {
            "_id": str(uuid.uuid4()),
            "spreadsheet_id": sheet_doc["_id"],
            "row_index": i,
            "cells": cells,
            "version": 1,
            "updated_by": None,
            "updated_at": now_iso()
        }
        bulk.append(row_doc)
        rows_created += 1
    if bulk:
        await db.rows.insert_many(bulk)
    return {"spreadsheet_id": sheet_doc["_id"], "rows": rows_created}

# Export XLSX
@app.get("/api/spreadsheets/{sheet_id}/export_xlsx")
async def export_xlsx(sheet_id: str):
    cursor = db.rows.find({"spreadsheet_id": sheet_id}).sort("row_index", 1)
    wb = openpyxl.Workbook()
    ws = wb.active
    # We will append rows in increasing row_index
    async for r in cursor:
        cells = r.get("cells", {})
        if not cells:
            ws.append([])
            continue
        # Find max column index for this row
        max_idx = max((openpyxl.utils.column_index_from_string(col) for col in cells.keys()), default=0)
        row_list = [None] * max_idx
        for col, c in cells.items():
            idx = openpyxl.utils.column_index_from_string(col)
            val = c.get("value") if isinstance(c, dict) else c
            row_list[idx - 1] = val
        ws.append(row_list)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                             headers={"Content-Disposition": f"attachment; filename=sheet_{sheet_id}.xlsx"})

@app.get("/ping")
async def ping():
    return {"ok": True}
